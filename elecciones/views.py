import csv
import hashlib
import io
import openpyxl
import qrcode
import base64
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.image import MIMEImage

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponse, FileResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import CandidaturaForm, EstudianteForm, VotoForm, ProcesoElectoralForm
from .models import Candidatura, Estudiante, Voto, ProcesoElectoral


def is_tribunal(user):
    return user.is_staff or user.is_superuser or user.groups.filter(name='Tribunal Electoral').exists()


def logout_view(request):
    """Manejo seguro de Logout para GET y POST sin error 405"""
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente.')
    return redirect('/accounts/login/')


@login_required
def dashboard(request):
    candidaturas = Candidatura.objects.all()
    estudiantes_count = Estudiante.objects.count()
    votos_count = Voto.objects.count()

    proceso = ProcesoElectoral.objects.first()
    if not proceso:
        proceso = ProcesoElectoral.objects.create(
            titulo="Elecciones Consejo Estudiantil",
            estado="Abierto"
        )

    # Conteo de votos por candidatura para gráficos en tiempo real
    conteo_votos = []
    total_votos = votos_count if votos_count > 0 else 1  # Prevenir división por cero
    
    colores_def = ['#4F46E5', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6', '#EC4899']
    for idx, cand in enumerate(candidaturas):
        v_count = Voto.objects.filter(eleccion=cand).count()
        porcentaje = round((v_count / total_votos) * 100, 2) if votos_count > 0 else 0
        color = cand.color_representativo or colores_def[idx % len(colores_def)]
        conteo_votos.append({
            'candidatura': cand,
            'votos': v_count,
            'porcentaje': porcentaje,
            'color': color,
        })

    # Verificar perfil de estudiante
    estudiante = getattr(request.user, 'estudiante_profile', None)

    return render(request, 'elecciones/dashboard.html', {
        'candidaturas': candidaturas,
        'estudiantes_count': estudiantes_count,
        'votos_count': votos_count,
        'conteo_votos': conteo_votos,
        'proceso': proceso,
        'estudiante': estudiante,
        'is_tribunal': is_tribunal(request.user),
    })


@login_required
def papeleta(request):
    proceso = ProcesoElectoral.objects.first()
    if not proceso:
        proceso = ProcesoElectoral.objects.create(
            titulo="Elecciones Consejo Estudiantil",
            estado="Abierto"
        )

    # 1. OBTENER O CREAR EL PERFIL DE ESTUDIANTE (Aplica para JOSE / Administradores y nuevos registros)
    estudiante = getattr(request.user, 'estudiante_profile', None)
    
    if not estudiante:
        # Crea un perfil de estudiante 'al vuelo' si el usuario autenticado es Admin / Staff o usuario nuevo sin perfil
        estudiante, _ = Estudiante.objects.get_or_create(
            usuario=request.user,
            defaults={
                'cedula_matricula': f"ADM-{request.user.id}",
                'grado_curso': 'Administración' if request.user.is_staff else 'Estudiante',
                'paralelo': 'A',
                'ha_votado': False
            }
        )

    # RESTRICCIÓN 1: Verificar si el Proceso Electoral está cerrado o fuera de horario
    if not proceso.is_activo():
        messages.error(request, 'El proceso electoral se encuentra actualmente CERRADO o fuera del horario establecido.')
        return render(request, 'elecciones/papeleta.html', {
            'bloqueado': True,
            'motivo_bloqueo': 'El proceso electoral está cerrado o fuera de horario.',
            'proceso': proceso,
        })

    # RESTRICCIÓN 2: Verificar estudiante y voto único (ha_votado == True)
    if estudiante and estudiante.ha_votado:
        messages.warning(request, 'Ya has ejercido tu derecho al voto. No puedes acceder a la papeleta nuevamente.')
        return render(request, 'elecciones/papeleta.html', {
            'bloqueado': True,
            'motivo_bloqueo': 'Ya has emitido tu voto en este proceso electoral.',
            'estudiante': estudiante,
            'proceso': proceso,
        })

    # Obtenemos todas las candidaturas registradas
    candidaturas = Candidatura.objects.all()

    if request.method == 'POST':
        if estudiante and estudiante.ha_votado:
            messages.error(request, 'Acción no permitida: Ya registraste un voto previo.')
            return redirect('elecciones:dashboard')

        cand_id = request.POST.get('candidatura_id')
        if not cand_id:
            messages.error(request, 'Por favor, selecciona una candidatura para emitir tu voto.')
        else:
            candidatura = get_object_or_404(Candidatura, pk=cand_id)
            
            # Generar hash único de seguridad (SHA-256)
            timestamp = timezone.now().timestamp()
            raw_hash_data = f"{request.user.id}-{candidatura.id}-{timestamp}-{settings.SECRET_KEY}"
            hash_seguridad = hashlib.sha256(raw_hash_data.encode('utf-8')).hexdigest()

            # 1. Registrar Voto (Garantizando voto secreto)
            voto = Voto.objects.create(
                eleccion=candidatura,
                fecha_hora=timezone.now(),
                hash_seguridad=hash_seguridad,
            )

            # 2. Marcar ha_votado = True en el perfil del Estudiante
            estudiante.ha_votado = True
            estudiante.save()

            # 3. Intentar envío de correo con QR
            try:
                qr_img = qrcode.make(f"VOTOESCOLAR-CERTIFICADO-DIGITAL:{hash_seguridad}")
                qr_io = io.BytesIO()
                qr_img.save(qr_io, format='PNG')
                qr_bytes = qr_io.getvalue()

                subject = "Certificado Digital de Votación - VOTOESCOLAR"
                body_html = f"""
                <div style="font-family: Arial, sans-serif; padding: 20px; color: #333; max-width: 600px; border: 2px solid #1E3A8A; border-radius: 12px; background: #ffffff;">
                    <div style="text-align: center; border-bottom: 2px solid #F5A425; padding-bottom: 15px; margin-bottom: 15px;">
                        <h2 style="color: #1E3A8A; margin: 0;">CERTIFICADO DIGITAL DE VOTACIÓN</h2>
                        <p style="color: #64748B; margin: 5px 0 0 0; font-size: 14px;">VOTOESCOLAR - Elecciones Estudiantiles</p>
                    </div>
                    <p>Estimado/a <strong>{request.user.get_full_name() or request.user.username}</strong>,</p>
                    <p>Se certifica que has emitido exitosamente tu voto en el Proceso Electoral Estudiantil de forma secreta y cifrada.</p>
                    <div style="background: #F8FAFC; padding: 15px; border-radius: 8px; border: 1px solid #E2E8F0; margin: 15px 0;">
                        <p style="margin: 3px 0;"><strong>Fecha y Hora:</strong> {voto.fecha_hora.strftime('%d/%m/%Y %H:%M:%S')}</p>
                        <p style="margin: 3px 0;"><strong>Hash de Transacción:</strong> <code style="background: #E2E8F0; padding: 2px 6px; border-radius: 4px;">{hash_seguridad}</code></p>
                    </div>
                </div>
                """

                destinatario = request.user.email if request.user.email else 'estudiante@votoescolar.edu.ec'

                msg = EmailMultiAlternatives(
                    subject=subject,
                    body=f"VOTOESCOLAR: Certificado digital de voto registrado. Hash: {hash_seguridad}",
                    from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@votoescolar.edu.ec'),
                    to=[destinatario],
                )
                msg.attach_alternative(body_html, "text/html")

                mime_img = MIMEImage(qr_bytes)
                mime_img.add_header('Content-ID', '<qr_code_image>')
                mime_img.add_header('Content-Disposition', 'inline', filename='certificado_qr.png')
                msg.attach(mime_img)

                msg.send(fail_silently=True)
            except Exception as e:
                print(f"Error al enviar correo con QR: {e}")

            messages.success(request, '¡Tu voto ha sido registrado con éxito!')
            return redirect('elecciones:confirmacion', pk=voto.pk)

    return render(request, 'elecciones/papeleta.html', {
        'candidaturas': candidaturas,
        'estudiante': estudiante,
        'proceso': proceso,
        'bloqueado': False,
    })


@login_required
def confirmacion(request, pk):
    voto = get_object_or_404(Voto, pk=pk)
    
    qr_img = qrcode.make(f"VOTOESCOLAR-CERTIFICADO-DIGITAL:{voto.hash_seguridad}")
    qr_io = io.BytesIO()
    qr_img.save(qr_io, format='PNG')
    qr_base64 = base64.b64encode(qr_io.getvalue()).decode('utf-8')

    return render(request, 'elecciones/confirmacion.html', {
        'voto': voto,
        'qr_base64': qr_base64,
    })


# --- EXPORTACIÓN E INFORMES ---

@login_required
@user_passes_test(is_tribunal)
def acta_final_pdf(request):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor('#1E3A8A')
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        alignment=1,
        textColor=colors.HexColor('#475569')
    )
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1E293B')
    )

    story.append(Paragraph("REPUBLICA DEL ECUADOR", subtitle_style))
    story.append(Paragraph("TRIBUNAL ELECTORAL ESTUDIANTIL", subtitle_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("ACTA OFICIAL DE ESCRUTINIO FINAL Y RESULTADOS", title_style))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1E3A8A'), spaceAfter=15))

    proceso = ProcesoElectoral.objects.first()
    fecha_str = timezone.now().strftime("%d de %B de %Y - %H:%M:%S")

    info_p = f"""
    <b>Proceso Electoral:</b> {proceso.titulo if proceso else 'Elecciones Consejo Estudiantil'}<br/>
    <b>Fecha de Emisión del Acta:</b> {fecha_str}<br/>
    <b>Estado del Proceso:</b> {proceso.estado if proceso else 'Finalizado'}
    """
    story.append(Paragraph(info_p, body_style))
    story.append(Spacer(1, 15))

    candidaturas = Candidatura.objects.all()
    total_votos = Voto.objects.count()

    table_data = [
        [Paragraph("<b>Lista / Candidatura</b>", body_style),
         Paragraph("<b>Presidente Candidato</b>", body_style),
         Paragraph("<b>Votos Obtenidos</b>", body_style),
         Paragraph("<b>Porcentaje</b>", body_style)]
    ]

    ganador = None
    max_v = -1

    for cand in candidaturas:
        v_count = Voto.objects.filter(eleccion=cand).count()
        pct = (v_count / total_votos * 100) if total_votos > 0 else 0
        if v_count > max_v:
            max_v = v_count
            ganador = cand
        table_data.append([
            Paragraph(cand.nombre_lista, body_style),
            Paragraph(cand.presidente, body_style),
            Paragraph(str(v_count), body_style),
            Paragraph(f"{pct:.2f}%", body_style),
        ])

    table_data.append([
        Paragraph("<b>TOTAL DE VOTOS EMITIDOS</b>", body_style),
        Paragraph("", body_style),
        Paragraph(f"<b>{total_votos}</b>", body_style),
        Paragraph("<b>100.00%</b>", body_style),
    ])

    t = Table(table_data, colWidths=[180, 160, 90, 90])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2E8F0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#94A3B8')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#64748B')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
    ]))

    story.append(t)
    story.append(Spacer(1, 20))

    if ganador and total_votos > 0:
        ganador_p = f"""
        <b>DECLARATORIA OFICIAL DE GANADOR:</b><br/>
        Concluido el escrutinio general de los sufragios, el Tribunal Electoral proclama formalmente a la Lista 
        <font color="#1E3A8A"><b>"{ganador.nombre_lista}"</b></font> (Presidente: <b>{ganador.presidente}</b>) 
        como la Ganadora del Proceso Electoral con <b>{max_v} votos</b> ({ (max_v/total_votos*100) if total_votos > 0 else 0:.2f}%).
        """
        story.append(Paragraph(ganador_p, body_style))
        story.append(Spacer(1, 25))

    story.append(Spacer(1, 40))
    firmas_data = [
        [Paragraph("_______________________________<br/><b>Presidente Tribunal Electoral</b>", body_style),
         Paragraph("_______________________________<br/><b>Secretario Tribunal Electoral</b>", body_style)]
    ]
    tf = Table(firmas_data, colWidths=[260, 260])
    tf.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(tf)

    doc.build(story)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='Acta_Final_Elecciones_VOTOESCOLAR.pdf')


@login_required
@user_passes_test(is_tribunal)
def export_padron_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Padrón Electoral"

    headers = ['N°', 'Cédula / Matrícula', 'Usuario', 'Nombre Completo', 'Email', 'Grado / Curso', 'Paralelo', 'Ha Votado', 'Fecha Registro']
    ws.append(headers)

    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    estudiantes = Estudiante.objects.select_related('usuario').all()
    
    for idx, est in enumerate(estudiantes, 1):
        nombre_comp = est.usuario.get_full_name() or est.usuario.username
        ha_votado_str = "SÍ" if est.ha_votado else "NO"
        fecha_reg = est.usuario.date_joined.strftime("%d/%m/%Y") if est.usuario.date_joined else "-"
        
        row = [
            idx,
            est.cedula_matricula,
            est.usuario.username,
            nombre_comp,
            est.usuario.email,
            est.grado_curso,
            est.paralelo,
            ha_votado_str,
            fecha_reg
        ]
        ws.append(row)
        current_row = idx + 1
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            if col_num in [1, 2, 6, 7, 8, 9]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            if col_num == 8:
                if est.ha_votado:
                    cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
                    cell.font = Font(color='15803D', bold=True)
                else:
                    cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    cell.font = Font(color='B91C1C', bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Padron_Electoral_Estudiantil_VOTOESCOLAR.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_tribunal)
def export_padron_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="padron_electoral_votoescolar.csv"'
    writer = csv.writer(response)
    writer.writerow(['Usuario', 'Nombre Completo', 'Email', 'Cédula/Matrícula', 'Grado/Curso', 'Paralelo', 'Ha Votado'])
    for est in Estudiante.objects.select_related('usuario').all():
        writer.writerow([
            est.usuario.username,
            est.usuario.get_full_name() or est.usuario.username,
            est.usuario.email,
            est.cedula_matricula,
            est.grado_curso,
            est.paralelo,
            'SÍ' if est.ha_votado else 'NO',
        ])
    return response


# --- VISTAS DE GESTIÓN (TRIBUNAL ELECTORAL) ---

@login_required
@user_passes_test(is_tribunal)
def candidatura_list(request):
    candidaturas = Candidatura.objects.all()
    return render(request, 'elecciones/candidatura_list.html', {'candidaturas': candidaturas})


@login_required
@user_passes_test(is_tribunal)
def candidatura_create(request):
    if request.method == 'POST':
        form = CandidaturaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidatura registrada correctamente.')
            return redirect('elecciones:candidatura_list')
    else:
        form = CandidaturaForm()
    return render(request, 'elecciones/candidatura_form.html', {'form': form, 'title': 'Nueva Candidatura'})


@login_required
@user_passes_test(is_tribunal)
def candidatura_update(request, pk):
    candidatura = get_object_or_404(Candidatura, pk=pk)
    if request.method == 'POST':
        form = CandidaturaForm(request.POST, request.FILES, instance=candidatura)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidatura actualizada correctamente.')
            return redirect('elecciones:candidatura_list')
    else:
        form = CandidaturaForm(instance=candidatura)
    return render(request, 'elecciones/candidatura_form.html', {'form': form, 'title': 'Editar Candidatura'})


@login_required
@user_passes_test(is_tribunal)
def candidatura_delete(request, pk):
    candidatura = get_object_or_404(Candidatura, pk=pk)
    if request.method == 'POST':
        candidatura.delete()
        messages.success(request, 'Candidatura eliminada.')
        return redirect('elecciones:candidatura_list')
    return render(request, 'elecciones/candidatura_confirm_delete.html', {'candidatura': candidatura})


@login_required
@user_passes_test(is_tribunal)
def estudiante_list(request):
    estudiantes = Estudiante.objects.select_related('usuario').all()
    return render(request, 'elecciones/estudiante_list.html', {'estudiantes': estudiantes})


@login_required
@user_passes_test(is_tribunal)
def estudiante_create(request):
    if request.method == 'POST':
        form = EstudianteForm(request.POST, request.FILES)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            first_name = form.cleaned_data.get('first_name', '')
            last_name = form.cleaned_data.get('last_name', '')
            email = form.cleaned_data.get('email', '')

            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=first_name,
                last_name=last_name,
                email=email
            )

            estudiante = form.save(commit=False)
            estudiante.usuario = user
            estudiante.save()

            messages.success(request, f'Estudiante "{user.username}" creado exitosamente.')
            return redirect('elecciones:estudiante_list')
    else:
        form = EstudianteForm()

    return render(request, 'elecciones/estudiante_form.html', {'form': form, 'title': 'Nuevo Estudiante'})


@login_required
@user_passes_test(is_tribunal)
def estudiante_update(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    user = estudiante.usuario

    if request.method == 'POST':
        form = EstudianteForm(request.POST, request.FILES, instance=estudiante)
        if form.is_valid():
            user.username = form.cleaned_data['username']
            user.first_name = form.cleaned_data.get('first_name', '')
            user.last_name = form.cleaned_data.get('last_name', '')
            user.email = form.cleaned_data.get('email', '')
            
            password = form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            user.save()

            form.save()
            messages.success(request, f'Estudiante "{user.username}" actualizado correctamente.')
            return redirect('elecciones:estudiante_list')
    else:
        initial_data = {
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
        }
        form = EstudianteForm(instance=estudiante, initial=initial_data)

    return render(request, 'elecciones/estudiante_form.html', {'form': form, 'title': 'Editar Estudiante'})


@login_required
@user_passes_test(is_tribunal)
def estudiante_delete(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)
    user = estudiante.usuario
    if request.method == 'POST':
        estudiante.delete()
        if user:
            user.delete()
        messages.success(request, 'Estudiante y su usuario fueron eliminados del padrón.')
        return redirect('elecciones:estudiante_list')
    return render(request, 'elecciones/estudiante_confirm_delete.html', {'estudiante': estudiante})


@login_required
@user_passes_test(is_tribunal)
def toggle_proceso(request):
    proceso = ProcesoElectoral.objects.first()
    if proceso:
        if proceso.estado == 'Abierto':
            proceso.estado = 'Cerrado'
            messages.warning(request, 'Proceso Electoral CERRADO.')
        else:
            proceso.estado = 'Abierto'
            messages.success(request, 'Proceso Electoral ABIERTO.')
        proceso.save()
    return redirect('elecciones:dashboard')


@login_required
@user_passes_test(is_tribunal)
def notificar_resultados_masivo(request):
    """Envía un correo masivo a todos los estudiantes con los resultados finales al cerrar el proceso."""
    candidaturas = Candidatura.objects.all()
    total_votos = Voto.objects.count()
    
    filas_resultados = ""
    for cand in candidaturas:
        v_count = Voto.objects.filter(eleccion=cand).count()
        pct = round((v_count / total_votos) * 100, 2) if total_votos > 0 else 0
        filas_resultados += f"""
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #E2E8F0;"><strong>{cand.nombre_lista}</strong> ({cand.presidente})</td>
            <td style="padding: 10px; border-bottom: 1px solid #E2E8F0; text-align: center;">{v_count} votos</td>
            <td style="padding: 10px; border-bottom: 1px solid #E2E8F0; text-align: right;"><strong>{pct}%</strong></td>
        </tr>
        """

    html_content = f"""
    <div style="font-family: Arial, sans-serif; padding: 20px; color: #333; max-width: 600px; border: 2px solid #1E3A8A; border-radius: 12px; background: #ffffff;">
        <div style="text-align: center; border-bottom: 2px solid #F5A425; padding-bottom: 15px; margin-bottom: 15px;">
            <h2 style="color: #1E3A8A; margin: 0;">RESULTADOS OFICIALES DE LAS ELECCIONES</h2>
            <p style="color: #64748B; margin: 5px 0 0 0; font-size: 14px;">Tribunal Electoral Estudiantil - VOTOESCOLAR</p>
        </div>
        <p>Estimada comunidad estudiantil,</p>
        <p>El proceso electoral ha sido formalmente <strong>CERRADO</strong>. A continuación, se detallan los resultados oficiales obtenidos:</p>
        
        <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
            <thead>
                <tr style="background: #1E3A8A; color: #ffffff;">
                    <th style="padding: 10px; text-align: left;">Candidatura / Lista</th>
                    <th style="padding: 10px; text-align: center;">Votos</th>
                    <th style="padding: 10px; text-align: right;">Porcentaje</th>
                </tr>
            </thead>
            <tbody>
                {filas_resultados}
            </tbody>
        </table>

        <div style="background: #F8FAFC; padding: 12px; border-radius: 8px; text-align: center; border: 1px solid #E2E8F0;">
            <p style="margin: 0; color: #475569; font-size: 13px;"><strong>Total de votos emitidos:</strong> {total_votos}</p>
        </div>
        <p style="margin-top: 20px; font-size: 12px; color: #94A3B8; text-align: center;">Este es un mensaje automático emitido por el sistema VOTOESCOLAR.</p>
    </div>
    """

    correos = list(Estudiante.objects.values_list('usuario__email', flat=True))
    correos_validos = [c for c in correos if c]

    if correos_validos:
        msg = EmailMultiAlternatives(
            subject="RESULTADOS OFICIALES - Elecciones Estudiantiles VOTOESCOLAR",
            body=f"Resultados de la elección. Votos totales: {total_votos}",
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@votoescolar.edu.ec'),
            bcc=correos_validos
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send(fail_silently=True)

        messages.success(request, f'Notificación masiva de resultados enviada con éxito a {len(correos_validos)} estudiantes.')
    else:
        messages.warning(request, 'No se encontraron correos de estudiantes válidos en el padrón.')

    return redirect('elecciones:dashboard')